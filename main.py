import requests
from unidecode import unidecode
import tkinter as tk
from tkinter import messagebox, scrolledtext, ttk
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime, timedelta
import pandas as pd
import os
import warnings
from openpyxl.styles import PatternFill, Border, Side, Font
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook

df_tennis = None
df_foot = None


def fetch_data_multi(url, headers):
    try:
        response = requests.get(url, headers=headers)
        if response.status_code == 200:
            print_to_app(f"Traitement du match en cours...")
            return url, response.json()  # return the URL as well to know which data belongs to which event
        else:
            print_to_app(f"Erreur HTTP {response.status_code}... [Pas de donnÃ©es odds pour ce match]")
            return url, None
    except requests.exceptions.ConnectionError as e:
        print_to_app(f"Une erreur s'est produite lors de la rÃ©cupÃ©ration des donnÃ©es du match.. [perdu]")
        return url, None

def fetch_team_ids(url, headers):
    try:
        response = requests.get(url, headers=headers)
        if response.status_code == 200:
            data = response.json()
            if 'event' in data:
                event = data['event']
                if 'homeTeam' in event and 'id' in event['homeTeam']:
                    home_team_id = event['homeTeam']['id']
                else:
                    home_team_id = None

                if 'awayTeam' in event and 'id' in event['awayTeam']:
                    away_team_id = event['awayTeam']['id']
                else:
                    away_team_id = None

                return home_team_id, away_team_id
        else:
            print_to_app(f"HTTP error {response.status_code}... [No data for this event]")
            return None, None
    except requests.exceptions.ConnectionError as e:
        print_to_app(f"An error occurred while fetching match data... [lost connection]")
        return None, None


def fetch_last_5_games(team_id, headers):
    url = f"https://api.sofascore.com/api/v1/team/{team_id}/events/last/0"
    response = requests.get(url, headers=headers)
    data = response.json()

    if "events" not in data:
        return None

    # Sort the events by startTimestamp
    sorted_games = sorted(data['events'], key=lambda x: x['startTimestamp'], reverse=True)

    # We only need the last 5 games
    games = sorted_games[:5]
    results = []

    for game in games:
        winner_code = game.get("winnerCode")
        home_team_id = game.get("homeTeam", {}).get("id")
        away_team_id = game.get("awayTeam", {}).get("id")

        if winner_code == 1 and team_id == home_team_id:
            results.append("W")
        elif winner_code == 2 and team_id == away_team_id:
            results.append("W")
        elif winner_code == 1 and team_id == away_team_id:
            results.append("L")
        elif winner_code == 2 and team_id == home_team_id:
            results.append("L")
        else:
            results.append("D")

    return " | ".join(results)

def extract_ids_and_tournaments(json_data, given_date):
    output = []

    # Convert the date string to a datetime object and set the time to the start of the day
    given_date = datetime.strptime(given_date, "%Y-%m-%d")
    given_date = given_date.replace(hour=0, minute=0, second=0)

    # Calculate the start and end dates for the range we're interested in
    start_date = given_date - timedelta(days=1)
    end_date = given_date + timedelta(days=1)

    # Get the current datetime
    current_datetime = datetime.now()

    # Iterate over each event
    for event in json_data['events']:
        if 'crowdsourcingDataDisplayEnabled' in event and 'id' in event:

            # Get the event's timestamp and convert it to a datetime object
            event_datetime = datetime.fromtimestamp(event['startTimestamp'])

            # Ignore events that don't occur within the date range or that have already occurred
            if not start_date <= event_datetime <= end_date or event_datetime <= current_datetime:
                continue

            # If the event is within the date range and has not yet occurred, add its id and tournament to the output
            data = {
                "id": event['id'],
                "uniqueTournament": event['tournament']['uniqueTournament']['name'],
                "time": event_datetime.strftime("%H:%M")
            }
            output.append(data)

    # Sort the output by time
    output.sort(key=lambda x: x['time'])

    return output


def extract_data(data, threshold_percentage, id_tournament_map, team_ids):
    output = []

    for id, id_data in data.items():
        clubs = [
            unidecode(id_data['markets'][-1]['choices'][0]['name']),
            unidecode(id_data['markets'][-1]['choices'][-1]['name'])
        ]

        if '1' in clubs or '2' in clubs:
            continue

        club_line = ' vs '.join(clubs)
        sport = 'Football'
        tournament = id_tournament_map[id][
            'uniqueTournament']  # Get the tournament for this id, or an empty string if not found

        max_value_details = []
        max_value = 0

        for market in id_data['markets']:
            if market['marketName'] == 'Full time':
                for choice in market['choices']:
                    if choice['name'] in ['1', 'X', '2']:
                        fractional = choice['fractionalValue'].split('/')
                        float_value = float(fractional[0]) / float(fractional[1]) + 1
                        probability = 1 / float_value
                        float_value = round(float_value, 2)
                        probability = round(probability, 2)

                        team_name = clubs[0] if choice['name'] == '1' else clubs[1]
                        analysis = f"Cotes: {float_value}"
                        winning_team_id = team_ids[id]['home'] if choice['name'] == '1' else team_ids[id]['away']

                        if probability > max_value:  # Update if this is a larger value
                            max_value = probability
                            max_value_details = [id_tournament_map[id]['time'], club_line, sport, tournament, analysis,
                                                 choice['name'], team_name, probability * 100,
                                                 winning_team_id]

        if max_value_details and max_value > threshold_percentage:  # Append only if probability is higher than the threshold
            output.append(max_value_details)

    return output


def print_to_app(message):
    console.insert(tk.END, message + '\n')
    console.see(tk.END)

def run_script():
    date = date_entry.get()
    probability = prob_entry.get()

    try:
        given_date = datetime.strptime(date, "%Y-%m-%d").date()
        current_date = datetime.now().date()
        if given_date < current_date:
            messagebox.showerror("Erreur", "La date saisie est passÃ©e. Veuillez saisir une date future.")
            return
        probability = float(probability)
        if probability < 0 or probability > 1:
            messagebox.showerror("Erreur", "La probabilitÃ© doit Ãªtre une valeur entre 0 et 1.")
            return
    except ValueError:
        messagebox.showerror("Erreur", "EntrÃ©e invalide. Veuillez saisir une date et une probabilitÃ© valides.")
        return

    # Disable the button while the script is running
    run_button.config(state="disabled")
    console.delete(1.0, tk.END)

    threading.Thread(target=foot, args=(date, probability), daemon=True).start()
    threading.Thread(target=tennis, args=(date, probability), daemon=True).start()

def extract_ids_tournaments_and_players(json_data, given_date):
    output = []

    # Convert the date string to a datetime object and set the time to the start of the day
    given_date = datetime.strptime(given_date, "%Y-%m-%d")
    given_date = given_date.replace(hour=0, minute=0, second=0)

    # Calculate the start and end dates for the range we're interested in
    start_date = given_date - timedelta(days=1)
    end_date = given_date + timedelta(days=1)

    # Get the current datetime
    current_datetime = datetime.now()

    # Iterate over each event
    for event in json_data['events']:
        if 'crowdsourcingDataDisplayEnabled' in event and 'id' in event:

            # Get the event's timestamp and convert it to a datetime object
            event_datetime = datetime.fromtimestamp(event['startTimestamp'])
            # Ignore events that don't occur within the date range or that have already occurred
            if not start_date <= event_datetime <= end_date or event_datetime <= current_datetime:
                continue
            # If the event is within the date range and has not yet occurred, add its id, tournament, players' short names, and timestamp to the output
            data = {
                "id": event['id'],
                "uniqueTournament": event['tournament']['uniqueTournament']['name'],
                "homePlayer": event['homeTeam']['shortName'],
                "awayPlayer": event['awayTeam']['shortName'],
                "time": event_datetime.strftime("%H:%M")  # Add time in HH:MM format
            }
            output.append(data)

    # Sort the output by time
    output.sort(key=lambda x: x['time'])

    return output


def extract_data_tennis(data, winning_odds_data, threshold_percentage, id_tournament_map,team_ids):
    output = []
    for id, id_data in data.items():
        players = [
            unidecode(id_tournament_map[id]['homePlayer']),
            unidecode(id_tournament_map[id]['awayPlayer'])
        ]

        if '1' in players or '2' in players:
            continue

        players_line = ' vs '.join(players)
        sport = 'Tennis'
        tournament = id_tournament_map[id]['uniqueTournament']

        if id in winning_odds_data:
            odds_data = winning_odds_data[id]
            actual_home = odds_data['home']['actual'] if odds_data['home'] is not None else 0
            actual_away = odds_data['away']['actual'] if odds_data['away'] is not None else 0
            actual = max(actual_home, actual_away)

            if actual > threshold_percentage * 100:
                max_value_details = []
                max_value = 0

                for market in id_data['markets']:
                    if market['marketName'] == 'Full time':
                        for choice in market['choices']:
                            if choice['name'] in ['1', '2']:
                                fractional = choice['fractionalValue'].split('/')
                                try:
                                    float_value = float(fractional[0]) / float(fractional[1]) + 1
                                except ValueError:
                                    print_to_app(f'Could not convert {fractional} to float')
                                except IndexError:
                                    print_to_app(f'Index error with {fractional}')
                                probability = 1 / float_value
                                float_value = round(float_value, 2)
                                probability = round(probability, 2)
                                player_name = players[0] if choice['name'] == '1' else players[1]
                                analysis = f"Cotes: {float_value}"
                                player_favorite = f"{player_name}"
                                winning_team_id = team_ids[id]['home'] if choice['name'] == '1' else team_ids[id]['away']
                                if probability > max_value:
                                    max_value = probability
                                    max_value_details = [id_tournament_map[id]['time'], players_line, sport, tournament, analysis,
                                                         choice['name'], player_favorite, probability * 100,
                                                         actual, winning_team_id]
                if max_value_details:
                    output.append(max_value_details)

    return output

import threading


warnings.simplefilter(action='ignore', category=FutureWarning)
write_lock = threading.Lock()

def save_to_sheet(df, sheet_name,lines):
    filename = 'Data.xlsx'
    if not filename.endswith(('.xlsx', '.xlsm', '.xls')):
        print_to_app(f"{filename} is not an Excel file.")
        return

    with write_lock:
        try:
            if os.path.exists(filename):  # If file exists, load existing workbook
                writer = pd.ExcelWriter(filename, engine='openpyxl', mode='a', if_sheet_exists='replace')
                writer.workbook = load_workbook(filename)
            else:  # If file does not exist, create new workbook
                writer = pd.ExcelWriter(filename, engine='openpyxl', mode='w')

            df.to_excel(writer, index=False, sheet_name=sheet_name)
            writer.close()
        except Exception as e:
            print_to_app(f"Error occurred while accessing {filename}: {str(e)}")
            return

    # Process the saved workbook and apply styles
    try:
        book = load_workbook(filename)
        sheet = book[sheet_name]
    except Exception as e:
        print_to_app(f"Error occurred while accessing {filename}: {str(e)}")
        return

    # Define the color fills
    redFill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    greenFill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
    yellowFill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
    headerFill = PatternFill(start_color='1E90FF', end_color='1E90FF', fill_type='solid')  # Light blue fill

    # Set the border style
    border = Border(bottom=Side(style='thick', color='ff5f87b3'))
    matches_border = Border(
        right=Side(style='thick', color='ff001ae5'),
        left=Side(style='thick', color='ff001ae5'),
        bottom=Side(style='thick', color='000000')
    )

    desired_height = 18
    desired_width = 3  # Adjust this value to set the width manually

    # Style the header row
    for cell in sheet[1]:
        cell.fill = headerFill
        cell.font = Font(bold=True, color='FFFFFF')  # White text

    # Find the 5Matches column index
    column_index = None
    for i, column in enumerate(sheet[1]):
        if column.value == "5Matches":
            column_index = i + 1
            break

    # Check if 5Matches column was found
    if column_index is None:
        print_to_app("Couldn't find 5Matches column.")
        return

    # Iterate over the columns
    for column in sheet.iter_cols():
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column[0].column_letter].width = adjusted_width

    # Extend the "5Matches" cell to span across the 5 newly created columns
    sheet.merge_cells(start_row=1, start_column=column_index, end_row=1, end_column=column_index + 4)

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        for cell in row:
            cell.border = border
        sheet.row_dimensions[row[0].row].height = desired_height

    # Iterate over the cells in the 5Matches column
    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row=row, column=column_index)
        results = cell.value.split(' | ')

        # Create a new column for each result
        for i, result in enumerate(results):
            new_cell = sheet.cell(row=row, column=column_index + i)
            new_cell.value = result
            if result == 'W':
                new_cell.fill = greenFill
            elif result == 'L':
                new_cell.fill = redFill
            elif result == 'D':
                new_cell.fill = yellowFill

            new_cell.border = matches_border

            # Set the height of the row to fit the contents
            sheet.row_dimensions[row].height = desired_height

            # Set the width of the new columns to fit the contents
            sheet.column_dimensions[get_column_letter(column_index + i)].width = desired_width

    # Save the workbook
    try:
        book.save(filename)
        print_to_app(f"TerminÃ©. Les rÃ©sultats sont disponibles dans le fichier '{filename}' sur la feuille '{sheet_name}'.")
    except Exception as e:
        print_to_app(f'Error closing the excel file: {str(e)}')

def tennis(date, threshold_probability):
    print_to_app("[Tennis] DÃ©but de l'opÃ©ration...")

    # Convert the date string to a datetime object
    given_date = datetime.strptime(date, "%Y-%m-%d")
    original_date_str = given_date.strftime("%Y-%m-%d")  # Store the original date string

    # Calculate the dates for the day before, the given date, and the day after
    dates = [given_date - timedelta(days=1), given_date, given_date + timedelta(days=1)]

    headers = {
        'User-Agent': 'Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:15.0) Gecko/20100101 Firefox/15.0.1',
    }

    all_data = {}
    winning_odds_data = {}
    id_tournament_map = {}
    team_ids = {}

    # Loop over the dates
    for date in dates:
        # Format the date as a string
        date_str = date.strftime("%Y-%m-%d")

        url = f'https://api.sofascore.com/api/v1/sport/tennis/scheduled-events/{date_str}'

        print_to_app("RÃ©cupÃ©ration des donnÃ©es ...")
        url, data = fetch_data_multi(url, headers)

        print_to_app("Extraction des Ã©vÃ©nements...")
        event_ids_tournaments = extract_ids_tournaments_and_players(data, original_date_str)

        # Instead of just the ids, now we have a list of dictionaries with ids and tournaments
        id_tournament_map.update({
            event['id']: {
                'uniqueTournament': event['uniqueTournament'],
                'homePlayer': event['homePlayer'],
                'awayPlayer': event['awayPlayer'],
                'time': event['time']  # add 'time' here
            }
            for event in event_ids_tournaments
        })


        temps_estime = len(id_tournament_map)

        if temps_estime < 60:
            temps_affiche = f"{temps_estime} secondes"
        else:
            minutes = temps_estime // 60
            secondes = temps_estime % 60
            temps_affiche = f"{minutes} minute{'s' if minutes > 1 else ''} et {secondes} seconde{'s' if secondes > 1 else ''}"

        print_to_app(f"RÃ©cupÃ©ration et analyse des donnÃ©es pour chaque Ã©vÃ©nement... [Temps estimÃ© : {temps_affiche}]")

        with ThreadPoolExecutor() as executor:
            futures = [executor.submit(fetch_team_ids, f"https://api.sofascore.com/api/v1/event/{id_value}", headers)
                       for id_value in id_tournament_map]
            for future, id_value in zip(futures, id_tournament_map):
                home_team_id, away_team_id = future.result()
                if home_team_id is not None and away_team_id is not None:
                    team_ids[id_value] = {"home": home_team_id, "away": away_team_id}

        with ThreadPoolExecutor() as executor:
            futures = [
                executor.submit(fetch_data_multi, f"https://api.sofascore.com/api/v1/event/{id_value}/odds/1/all",
                                headers) for id_value in id_tournament_map]
            for future in futures:
                url, api_data = future.result()
                if api_data is not None:
                    id_value = int(url.split('/')[-4])
                    all_data[id_value] = api_data

        with ThreadPoolExecutor() as executor:
            futures = [executor.submit(fetch_data_multi,
                                       f"https://api.sofascore.com/api/v1/event/{id_value}/provider/1/winning-odds",
                                       headers) for id_value in id_tournament_map]
            for future in futures:
                url, odds_data = future.result()
                if odds_data is not None:
                    id_value = int(url.split('/')[-4])
                    winning_odds_data[id_value] = odds_data

    print_to_app("Extraction des donnÃ©es finales...")
    lines = extract_data_tennis(all_data, winning_odds_data, threshold_probability,
                         id_tournament_map,team_ids)

    with ThreadPoolExecutor() as executor:
        futures = {line[-1]: executor.submit(fetch_last_5_games, line[-1], headers)
                   for line in lines}

    team_results = {team_id: future.result() for team_id, future in futures.items()}

    # Replace team_id in lines with the fetched results
    for line in lines:
        team_id = line[-1]
        line[-1] = team_results[team_id]  # Fetch the last 5 games results and replace the team_id with the results
    print_to_app("CrÃ©ation du fichier Excel...")
    df = pd.DataFrame(lines, columns=['Time', 'Ã‰quipes', 'Sport', 'Ligue', 'Analyse', 'Choix', 'Ã‰quipe favorite',
                                      'ProbabilitÃ© ThÃ©orique', '5Matches'])
    save_to_sheet(df, 'Tennis',lines)
    print_to_app(f"TerminÃ©. Les rÃ©sultats sont disponibles dans le fichier Data sur la feuille Tennis.")
    run_button.config(state="normal")


def foot(date, threshold_probability):
    print_to_app("[Football] DÃ©but de l'opÃ©ration... ")

    # Convert the date string to a datetime object
    given_date = datetime.strptime(date, "%Y-%m-%d")
    original_date_str = given_date.strftime("%Y-%m-%d")  # Store the original date string

    # Calculate the dates for the day before, the given date, and the day after
    dates = [given_date - timedelta(days=1), given_date, given_date + timedelta(days=1)]
    headers = {
        'User-Agent': 'Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:15.0) Gecko/20100101 Firefox/15.0.1',
    }

    all_data = {}
    winning_odds_data = {}
    id_tournament_map = {}
    team_ids = {}

    # Loop over the dates
    for date in dates:
        # Format the date as a string
        date_str = date.strftime("%Y-%m-%d")

        url = f'https://api.sofascore.com/api/v1/sport/football/scheduled-events/{date_str}'

        print_to_app("RÃ©cupÃ©ration des donnÃ©es ...")
        url , data = fetch_data_multi(url, headers)

        print_to_app("Extraction des Ã©vÃ©nements...")
        event_ids_tournaments = extract_ids_and_tournaments(data, original_date_str)

        # Instead of just the ids, now we have a list of dictionaries with ids and tournaments
        id_tournament_map.update({
            event['id']: {
                'uniqueTournament': event['uniqueTournament'],
                'time': event['time']  # add 'time' here
            }
            for event in event_ids_tournaments
        })
        temps_estime = len(id_tournament_map)

        if temps_estime < 60:
            temps_affiche = f"{temps_estime} secondes"
        else:
            minutes = temps_estime // 60
            secondes = temps_estime % 60
            temps_affiche = f"{minutes} minute{'s' if minutes > 1 else ''} et {secondes} seconde{'s' if secondes > 1 else ''}"


        print_to_app(f"RÃ©cupÃ©ration et analyse des donnÃ©es pour chaque Ã©vÃ©nement... [Temps estimÃ© : {temps_affiche}]")

        with ThreadPoolExecutor() as executor:
            futures = [executor.submit(fetch_team_ids, f"https://api.sofascore.com/api/v1/event/{id_value}", headers)
                       for id_value in id_tournament_map]
            for future, id_value in zip(futures, id_tournament_map):
                home_team_id, away_team_id = future.result()
                if home_team_id is not None and away_team_id is not None:
                    team_ids[id_value] = {"home": home_team_id, "away": away_team_id}

        with ThreadPoolExecutor() as executor:
            futures = [executor.submit(fetch_data_multi, f"https://api.sofascore.com/api/v1/event/{id_value}/odds/1/all", headers) for id_value in id_tournament_map]
            for future in futures:
                url, api_data = future.result()
                if api_data is not None:
                    id_value = int(url.split('/')[-4])
                    all_data[id_value] = api_data

        with ThreadPoolExecutor() as executor:
            futures = [executor.submit(fetch_data_multi, f"https://api.sofascore.com/api/v1/event/{id_value}/provider/1/winning-odds", headers) for id_value in id_tournament_map]
            for future in futures:
                url, odds_data = future.result()
                if odds_data is not None:
                    id_value = int(url.split('/')[-4])
                    winning_odds_data[id_value] = odds_data

    print_to_app("Extraction des donnÃ©es finales...")
    lines = extract_data(all_data, winning_odds_data, threshold_probability,
                         id_tournament_map,team_ids)

    with ThreadPoolExecutor() as executor:
        futures = {line[-1]: executor.submit(fetch_last_5_games, line[-1], headers)
                   for line in lines}

    team_results = {team_id: future.result() for team_id, future in futures.items()}

    # Replace team_id in lines with the fetched results
    for line in lines:
        team_id = line[-1]
        line[-1] = team_results[team_id]  # Fetch the last 5 games results and replace the team_id with the results
    print_to_app("CrÃ©ation du fichier Excel...")
    # assuming df_foot and df_tennis are your dataframes
    df = pd.DataFrame(lines, columns=['Time', 'Ã‰quipes', 'Sport', 'Ligue', 'Analyse', 'Choix', 'Ã‰quipe favorite',
                                      'ProbabilitÃ© ThÃ©orique', 'ProbabilitÃ©', '5Matches'])
    save_to_sheet(df, 'Football',lines)
    print_to_app(f"TerminÃ©. Les rÃ©sultats sont disponibles dans le fichier Data sur la feuille Football.")

    run_button.config(state="normal")


root = tk.Tk()
root.title("Assistant d'extraction de données")

largeur_fenetre = 500
hauteur_fenetre = 400
root.geometry(f'{largeur_fenetre}x{hauteur_fenetre}')

main_frame = ttk.Frame(root, padding="20 20 20 20")
main_frame.grid(column=0, row=0, sticky=(tk.W, tk.E, tk.N, tk.S))

root.columnconfigure(0, weight=1)
root.rowconfigure(0, weight=1)

sports_var = tk.StringVar(root)

date_label = ttk.Label(main_frame, text="Entrez la date (AAAA-MM-JJ) :", font=("Arial", 10))
date_label.grid(column=0, row=1, sticky=tk.W, padx=10, pady=10)

date_entry = ttk.Entry(main_frame, width=30)
date_entry.grid(column=1, row=1, sticky=(tk.W, tk.E), padx=10, pady=10)

prob_label = ttk.Label(main_frame, text="Entrez la probabilité (entre 0 et 1) :", font=("Arial", 10))
prob_label.grid(column=0, row=2, sticky=tk.W, padx=10, pady=10)

prob_entry = ttk.Entry(main_frame, width=30)
prob_entry.grid(column=1, row=2, sticky=(tk.W, tk.E), padx=10, pady=10)

run_button = ttk.Button(main_frame, text="Démarrer l'extraction", command=run_script)
run_button.grid(column=1, row=3, sticky=tk.E, padx=10, pady=20)

console = scrolledtext.ScrolledText(main_frame, width=40, height=15)
console.grid(column=0, row=4, columnspan=2, sticky=(tk.W, tk.E, tk.N, tk.S), padx=10, pady=10)

root.mainloop()
